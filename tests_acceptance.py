"""
Boutique Manager - acceptance tests (PRD section 12).

Runs against a throwaway SQLite DB (BOUTIQUE_DB) using Flask's test client.
Every acceptance criterion is exercised; failures print and exit non-zero.

Run:  python tests_acceptance.py
"""

import base64
import io
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# --- Point the app at a throwaway DB BEFORE importing app.py ---
_TMP_DB = tempfile.mktemp(suffix=".db", prefix="boutique_test_")
if os.path.exists(_TMP_DB):  # start clean even if a stale temp DB lingers
    os.remove(_TMP_DB)
os.environ["BOUTIQUE_DB"] = _TMP_DB

from app import create_app, kpi_totals  # noqa: E402
from models import db, Product, Sale, Expense, User  # noqa: E402

PASS = []
FAIL = []


def check(label, condition):
    (PASS if condition else FAIL).append(label)
    print(("  OK  " if condition else "ERREUR ") + label)


# A real 1x1 PNG (smallest valid) for the upload test.
PNG_1PX = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


def new_app():
    return create_app()


def main():
    app = new_app()
    client = app.test_client()

    # The app now requires authentication: create/refresh the default admin + sign in.
    with app.app_context():
        admin = User.query.filter_by(username="admin").first()
        if admin is None:
            admin = User(username="admin")
            db.session.add(admin)
        admin.set_password("admin123")
        db.session.commit()
    r = client.post("/login", data={"username": "admin", "password": "admin123"},
                    follow_redirects=True)
    check("Connexion de test reussie", r.status_code == 200)

    print("\n== 1. Pages de base ==")
    for url, label in [
        ("/", "Tableau de bord"),
        ("/products", "Liste produits"),
        ("/sales", "Historique ventes"),
        ("/sales/new", "Formulaire vente"),
        ("/expenses", "Liste depenses"),
        ("/expenses/new", "Formulaire depense"),
        ("/export", "Page export"),
    ]:
        r = client.get(url)
        check(f"GET {url} -> 200 ({label})", r.status_code == 200)

    with app.app_context():
        check("Base vide au demarrage (aucun produit)", Product.query.count() == 0)

    print("\n== 2. Ajout produit avec image ==")
    r = client.post(
        "/products/new",
        data={
            "name": "Test A",
            "category": "Electronique",
            "purchase_price": "4.00",
            "sale_price": "6.00",
            "low_stock_threshold": "1",
            "initial_stock": "10",
            "image": (io.BytesIO(PNG_1PX), "photo.png"),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    check("POST /products/new -> 200", r.status_code == 200)
    with app.app_context():
        p = Product.query.filter_by(name="Test A").first()
        check("Produit cree avec stock initial 10", p is not None and p.stock == 10)
        check("Image enregistree", p.image_url is not None)
        # Image is stored on ImageKit CDN; we cannot verify a local file.

    r = client.get("/products")
    check("Produit visible dans la liste", "Test A" in r.get_data(as_text=True))
    # Image lives on the ImageKit CDN (v2) - no local file exists to check.

    print("\n== 3. Vente 2 u. au prix liste ==")
    with app.app_context():
        pid = Product.query.filter_by(name="Test A").first().id
    client.post(
        "/sales/new",
        data={"product_id": str(pid), "quantity": "2", "charged_price": "6.00"},
        follow_redirects=True,
    )
    with app.app_context():
        p = Product.query.filter_by(name="Test A").first()
        s = Sale.query.filter_by(product_id=pid).order_by(Sale.id.desc()).first()
        check("Stock 10 - 2 = 8", p.stock == 8)
        check("Vente creee (liste 6, facture 6, remise 0)", s.total == 12.0 and s.discount_per_unit == 0.0)
        t = kpi_totals()
        check("KPI ventes = 12,00", abs(t["total_sales"] - 12.0) < 0.01)
        check("KPI profit = 4,00 (12 - 4*2)", abs(t["total_profit"] - 4.0) < 0.01)

    print("\n== 4. Vente remisee : liste 6,00 facture 5,00 x2 ==")
    client.post(
        "/sales/new",
        data={"product_id": str(pid), "quantity": "2", "charged_price": "5.00"},
        follow_redirects=True,
    )
    with app.app_context():
        p = Product.query.filter_by(name="Test A").first()
        s = Sale.query.filter_by(product_id=pid).order_by(Sale.id.desc()).first()
        check("Stock 8 - 2 = 6", p.stock == 6)
        check("Revenue reel = 10,00 (pas 12,00)", s.total == 10.0)
        check("Remise = 1,00 / unite", s.discount_per_unit == 1.0)
        check("Remise % = 16,67", abs(s.discount_percent - 16.67) < 0.01)
        t = kpi_totals()
        check("KPI ventes = 22,00", abs(t["total_sales"] - 22.0) < 0.01)
        check("KPI remises = 2,00", abs(t["total_discounts"] - 2.0) < 0.01)
        check("KPI profit = 6,00", abs(t["total_profit"] - 6.0) < 0.01)
        check("KPI stock = 6", t["units_stock"] == 6)

    print("\n== 5. Vente > stock bloquee ==")
    r = client.post(
        "/sales/new",
        data={"product_id": str(pid), "quantity": "99", "charged_price": "5.00"},
        follow_redirects=True,
    )
    body = r.get_data(as_text=True)
    with app.app_context():
        s = Sale.query.filter_by(product_id=pid).count()
        p = Product.query.filter_by(name="Test A").first()
        check("Message d'erreur FR affiche", "Stock insuffisant" in body)
        check("Aucune vente creee", s == 2)
        check("Stock inchange (6)", p.stock == 6)

    print("\n== 6. Remise negative / au-dessus du prix liste ==")
    r = client.post(
        "/sales/new",
        data={"product_id": str(pid), "quantity": "1", "charged_price": "7.00"},
        follow_redirects=True,
    )
    check("Prix facture > prix liste refuse", "ne peut pas dépasser" in r.get_data(as_text=True))
    r = client.post(
        "/sales/new",
        data={"product_id": str(pid), "quantity": "1", "charged_price": "-1"},
        follow_redirects=True,
    )
    check("Prix facture < 0 refuse", "Prix facturé invalide" in r.get_data(as_text=True))

    print("\n== 7. Remboursement (void) ==")
    with app.app_context():
        first_sale = Sale.query.order_by(Sale.id).first()
    client.post(f"/sales/{first_sale.id}/void", follow_redirects=True)
    with app.app_context():
        p = Product.query.filter_by(name="Test A").first()
        first_sale = db.session.get(Sale, first_sale.id)
        t = kpi_totals()
        check("Stock restaure (6 + 2 = 8)", p.stock == 8)
        check("Vente marquee remboursee", first_sale.refunded is True)
        check("KPI ventes excluent la vente (10,00)", abs(t["total_sales"] - 10.0) < 0.01)
        check("KPI remises exclues (2,00)", abs(t["total_discounts"] - 2.0) < 0.01)

    print("\n== 8. Statuts stock bas / rupture ==")
    client.post(
        "/products/new",
        data={"name": "Test B", "purchase_price": "1.00", "sale_price": "2.00",
              "low_stock_threshold": "1", "initial_stock": "1"},
        follow_redirects=True,
    )
    client.post(
        "/products/new",
        data={"name": "Test C", "purchase_price": "1.00", "sale_price": "2.00",
              "low_stock_threshold": "1", "initial_stock": "0"},
        follow_redirects=True,
    )
    with app.app_context():
        b = Product.query.filter_by(name="Test B").first()
        c = Product.query.filter_by(name="Test C").first()
        check("Test B statut 'low'", b.status == "low")
        check("Test C statut 'out' (rupture)", c.status == "out")
    r = client.get("/")
    check("Alerte stock bas sur tableau de bord", "Test B" in r.get_data(as_text=True))
    r = client.get("/products?status=out")
    check("Filtre rupture -> Test C visible", "Test C" in r.get_data(as_text=True))

    print("\n== 9. Depenses (frais) ==")
    r = client.post(
        "/expenses/new",
        data={"date": "2026-08-07", "description": "Livraison test", "amount": "10.50"},
        follow_redirects=True,
    )
    check("Depense creee", "Livraison test" in r.get_data(as_text=True))
    with app.app_context():
        t = kpi_totals()
        check("Frais KPI = 10,50", abs(t["operating"] - 10.5) < 0.01)
        # 40,00 (achat A 10x4) + 1,00 (achat B 1x1) + 10,50 (frais) = 51,50
        check("Depenses totales incluent la depense", abs(t["total_expenses"] - 51.5) < 0.01)
        check("La depense n'est PAS un produit", Product.query.filter_by(name="Livraison test").count() == 0)
    body = client.get("/sales/new").get_data(as_text=True)
    check("Depense absente du selecteur de vente", "Livraison test" not in body)
    r = client.get("/products")
    check("Depense absente de la liste produits", "Livraison test" not in r.get_data(as_text=True))

    print("\n== 10. Suppression douce ==")
    with app.app_context():
        cid = Product.query.filter_by(name="Test C").first().id
    client.post(f"/products/{cid}/delete", follow_redirects=True)
    r = client.get("/products")
    check("Produit masque de la liste active", "Test C" not in r.get_data(as_text=True))
    body = client.get("/sales/new").get_data(as_text=True)
    check("Produit absent du selecteur de vente", "Test C" not in body)
    with app.app_context():
        c = db.session.get(Product, cid)
        check("Produit conserve en base (soft-delete)", c is not None and c.is_deleted)

    print("\n== 11. Historique ventes (snapshots) ==")
    r = client.get("/sales")
    body = r.get_data(as_text=True)
    check("Vente remboursee listee comme telle", "Remboursée" in body)
    check("Remise % affichee (16,67 %)", "16.67" in body or "16,67" in body)

    print("\n== 12. Export ==")
    r = client.get("/export/xlsx")
    check("GET /export/xlsx -> 200", r.status_code == 200)
    if r.status_code == 200:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        check("Classeur avec 3 feuilles", wb.sheetnames == ["Feuille 1", "Ventes", "Dépenses"])
        ws = wb["Feuille 1"]
        check("Feuille 1 avec ligne Totaux", any(str(c.value) == "Totaux" for row in ws.iter_rows() for c in row))
        ws2 = wb["Ventes"]
        check("Feuille Ventes a une colonne Remise", "Remise / unité" in [c.value for c in ws2[1]])
    r = client.get("/export/csv/feuille1")
    check("GET /export/csv/feuille1 -> 200 (Totaux)", r.status_code == 200 and b"Totaux" in r.data)

    print("\n== 13. Persistance (redemarrage) ==")
    app2 = new_app()
    c2 = app2.test_client()
    c2.post("/login", data={"username": "admin", "password": "admin123"},
            follow_redirects=True)
    with app2.app_context():
        check("Produits toujours presents", Product.query.filter_by(name="Test A").count() == 1)
        t = kpi_totals()
        check("Totaux conserves (ventes 10,00)", abs(t["total_sales"] - 10.0) < 0.01)
        check("Depenses conservees (51,50)", abs(t["total_expenses"] - 51.5) < 0.01)
    r = c2.get("/")
    check("Dashboard ok apres redemarrage", r.status_code == 200)

    print("\n== 14. Pages d'erreur ==")
    r = client.get("/nope")
    check("404 convivial", r.status_code == 404 and "Page introuvable" in r.get_data(as_text=True))

    print("\n== 15. Langue : English ==")
    r = client.get("/lang/en", follow_redirects=True)
    check("Route /lang/en -> 200", r.status_code == 200)
    r = client.get("/")
    body = r.get_data(as_text=True)
    check("Menu anglais (Dashboard)", "Dashboard" in body)
    check("Menu anglais (Products)", "Products" in body)
    check("Devise format anglais (€10.00)", "€10.00" in body)
    check("Produits page en anglais", "/products" in body)
    r = client.get("/sales")
    check("Historique ventes en anglais (Sales history)", "Sales history" in r.get_data(as_text=True))
    r = client.get("/products")
    check("Ajouter un produit en anglais (Add a product)", "Add a product" in r.get_data(as_text=True))
    r = client.get("/login?lang=en")
    lb = r.get_data(as_text=True)
    check("Page de connexion en anglais (Sign in)", "Sign in" in lb)
    check("Page de connexion : champs anglais (Username)", "Username" in lb)

    print("\n== 16. Retour francais ==")
    client.get("/lang/fr", follow_redirects=True)
    body = client.get("/").get_data(as_text=True)
    check("Menu francais restaure (Tableau de bord)", "Tableau de bord" in body)

    print("\n== 17. Description produit (texte de publication) ==")
    styled = "✨ Washing Parfum\n- vrac de gouttes suffisent\n- 🌿 Lotus Dream"
    client.post(
        "/products/new",
        data={"name": "Test Desc", "purchase_price": "1.00", "sale_price": "2.00",
              "initial_stock": "2", "description": styled},
        follow_redirects=True,
    )
    with app.app_context():
        p = Product.query.filter_by(name="Test Desc").first()
        check("Description stockee verbatim (emojis + sauts)", p is not None and p.description == styled)
        dpid = p.id if p else 0
    body = client.get("/products").get_data(as_text=True)
    check("Description visible sur la liste", "Lotus Dream" in body or "🌿" in body)
    r = client.get("/products?q=Lotus")
    check("Recherche dans la description", "Test Desc" in r.get_data(as_text=True))
    r = client.get(f"/products/{dpid}/edit")
    eb = r.get_data(as_text=True)
    check("Formulaire pre-rempli avec la description", "Lotus" in eb)
    check("Bouton copier present", "Copier" in eb)

    # Cleanup temp files.
    try:
        os.remove(_TMP_DB)
        for f in os.listdir(os.path.join("static", "uploads")):
            os.remove(os.path.join("static", "uploads", f))
    except OSError:
        pass

    print("\n" + "=" * 50)
    print(f"  RESULTAT : {len(PASS)} reussis, {len(FAIL)} echecs")
    if FAIL:
        print("  ECHECS :")
        for f in FAIL:
            print("   - " + f)
    print("=" * 50)
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()